# v2 
import os
import re
import json
import random
import string
import asyncio
import base64
import tempfile
from datetime import datetime
from pathlib import Path

from telegram import Update, Bot, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes, ConversationHandler, CallbackQueryHandler
)
import anthropic
from playwright.async_api import async_playwright
from pypdf import PdfWriter, PdfReader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import gspread
from google.oauth2.service_account import Credentials

# ============================================================
# CONFIG
# ============================================================
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "").strip()
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "").strip()

if not TELEGRAM_TOKEN:
    raise ValueError("TELEGRAM_TOKEN not set!")
if not ANTHROPIC_API_KEY:
    raise ValueError("ANTHROPIC_API_KEY not set!")

# ============================================================
# GOOGLE SHEETS CONFIG
# ============================================================
GOOGLE_CREDENTIALS_FILE = os.environ.get("GOOGLE_CREDENTIALS_FILE", "google_credentials.json")
IRANFLY_SHEET_ID   = os.environ.get("IRANFLY_SHEET_ID", "")
FLYTICKET_SHEET_ID = os.environ.get("FLYTICKET_SHEET_ID", "")

GSHEETS_SCOPE = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ستون‌های گوگل شیت
COL_ROW        = 1
COL_REF        = 2
COL_PASSENGER  = 3
COL_ROUTE_OUT  = 4
COL_ROUTE_RET  = 5
COL_AIR_OUT    = 6
COL_AIR_RET    = 7
COL_DATE_OUT   = 8
COL_DATE_RET   = 9
COL_BUY        = 10
COL_SELL       = 11
COL_VOUCHER    = 12
COL_INSURANCE  = 13
COL_PROFIT     = 14
COL_PASSPORT   = 15
COL_CANCELLED  = 16

EXCEL_FILE             = "IRANFLY_tickets.xlsx"
TEMPLATE_FILE          = "IRANFLY_TEMPLATE_v1.html"
FLYTICKET_TEMPLATE_FILE = "FLYTICKET_TEMPLATE_v1.html"
PDF_PASSWORD           = "IRANFLY2025"

FLIGHTS = {
    "qeshm_ika_tbs": {"airline": "Qeshm Airline", "flight": "Q.2273",  "time": "11:00", "from": "IKA", "to": "TBS"},
    "varesh_ika_tbs": {"airline": "Varesh Airline","flight": "VR.6808", "time": "11:00", "from": "IKA", "to": "TBS"},
    "qeshm_tbs_ika": {"airline": "Qeshm Airline", "flight": "Q.2272",  "time": "14:50", "from": "TBS", "to": "IKA"},
    "varesh_tbs_ika": {"airline": "Varesh Airline","flight": "VR.6809", "time": "14:15", "from": "TBS", "to": "IKA"},
}

BAGGAGE = {
    "qeshm_economy":  "30 kg",
    "qeshm_business": "40 kg",
    "varesh_economy": "20 kg",
    "varesh_business":"20 kg",
}

# conversation states
WAITING_PASSPORT    = 1
WAITING_FLIGHT_INFO = 2
WAITING_REF_INPUT   = 3
WAITING_EDIT_VALUE  = 4
WAITING_CANCEL_CONFIRM = 5

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ============================================================
# HELPERS
# ============================================================
def gen_ref():
    while True:
        ref = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
        if any(c.isalpha() for c in ref) and any(c.isdigit() for c in ref):
            return ref

def extract_passport_info(image_bytes: bytes) -> dict:
    b64 = base64.b64encode(image_bytes).decode()
    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=500,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
            {"type": "text", "text": 'Extract from this passport: first_name, last_name, passport_number. Return ONLY JSON: {"first_name": "JOHN", "last_name": "DOE", "passport_number": "A12345678"}'}
        ]}]
    )
    text = re.sub(r'```json|```', '', response.content[0].text.strip()).strip()
    return json.loads(text)

def to_num(s):
    s = str(s).strip().replace(',', '.')
    if '.' in s:
        return int(float(s) * 1000)
    return int(s)

def parse_date(s):
    months = {"jan":"Jan","feb":"Feb","mar":"Mar","apr":"Apr","may":"May","jun":"Jun",
              "jul":"Jul","aug":"Aug","sep":"Sep","oct":"Oct","nov":"Nov","dec":"Dec"}
    s = s.lower().strip()
    m = re.match(r'(\d{1,2})\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', s)
    if m:
        return f"{m.group(1)} {months[m.group(2)]} 2026"
    return s

def parse_line(line):
    parts = [p.strip() for p in line.lower().split('/')]
    info = {
        "src": parts[0].upper() if len(parts) > 0 else "",
        "dst": parts[1].upper() if len(parts) > 1 else "",
        "airline": parts[2] if len(parts) > 2 else "varesh",
        "date": parse_date(parts[3]) if len(parts) > 3 else "",
        "sell": 0, "buy": 0, "voucher": 0, "insurance": 0,
        "flight_class": "business" if "business" in line.lower() else "economy"
    }
    for p in parts[4:]:
        p = p.strip()
        if p.startswith("sell"):
            info["sell"] = to_num(p[4:])
        elif p.startswith("buy"):
            info["buy"] = to_num(p[3:])
        elif p.startswith("voucher"):
            info["voucher"] = to_num(p[7:])
        elif p.startswith("insurance"):
            info["insurance"] = to_num(p[9:])
    return info

def parse_flight_input(text: str) -> dict:
    lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
    result = {
        "type": "oneway", "outbound": None, "return_flight": None,
        "sell_out": 0, "buy_out": 0, "sell_ret": 0, "buy_ret": 0,
        "voucher": 0, "insurance": 0, "class": "economy",
        "date_out": "", "date_ret": ""
    }
    if len(lines) >= 1:
        out = parse_line(lines[0])
        airline = "qeshm" if "qeshm" in out["airline"] else "varesh"
        key = f"{airline}_ika_tbs" if out["src"] == "IKA" else f"{airline}_tbs_ika"
        result["outbound"]   = FLIGHTS[key]
        result["sell_out"]   = out["sell"]
        result["buy_out"]    = out["buy"]
        result["voucher"]    = out["voucher"]
        result["insurance"]  = out["insurance"]
        result["class"]      = out["flight_class"]
        result["date_out"]   = out["date"]
    if len(lines) >= 2:
        result["type"] = "roundtrip"
        ret = parse_line(lines[1])
        airline_ret = "qeshm" if "qeshm" in ret["airline"] else "varesh"
        key_ret = f"{airline_ret}_ika_tbs" if ret["src"] == "IKA" else f"{airline_ret}_tbs_ika"
        result["return_flight"] = FLIGHTS[key_ret]
        result["sell_ret"]  = ret["sell"]
        result["buy_ret"]   = ret["buy"]
        result["date_ret"]  = ret["date"]
    return result

# ============================================================
# GOOGLE SHEETS
# ============================================================
def get_gsheet_ws(company: str):
    sheet_id = IRANFLY_SHEET_ID if company == "iranfly" else FLYTICKET_SHEET_ID
    if not sheet_id:
        return None
    creds = Credentials.from_service_account_file(GOOGLE_CREDENTIALS_FILE, scopes=GSHEETS_SCOPE)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    try:
        return sh.worksheet("بلیت‌ها")
    except gspread.WorksheetNotFound:
        return sh.get_worksheet(0)

def update_google_sheet(company: str, passenger: dict, flight: dict, ref_out: str, ref_ret: str):
    try:
        ws = get_gsheet_ws(company)
        if not ws:
            print(f"[WARN] Sheet ID for {company} not set.")
            return

        outbound  = flight.get("outbound") or {}
        return_fl = flight.get("return_flight") or {}
        is_round  = flight.get("type") == "roundtrip"
        total_sell = flight.get("sell_out", 0) + (flight.get("sell_ret", 0) if is_round else 0)
        total_buy  = flight.get("buy_out",  0) + (flight.get("buy_ret",  0) if is_round else 0)
        profit     = (total_sell - total_buy) + flight.get("voucher", 0) + flight.get("insurance", 0)
        next_row   = len(ws.get_all_values()) + 1

        new_row = [
            next_row - 1,
            ref_out,
            f"{passenger['first_name']} {passenger['last_name']}",
            f"{outbound.get('from','-')} → {outbound.get('to','-')}",
            f"{return_fl.get('from','-')} → {return_fl.get('to','-')}" if is_round else "-",
            f"{outbound.get('airline','-')} {outbound.get('flight','')}",
            f"{return_fl.get('airline','-')} {return_fl.get('flight','')}" if is_round else "-",
            flight.get("date_out", "-"),
            flight.get("date_ret", "-") if is_round else "-",
            total_buy,
            total_sell,
            flight.get("voucher", 0),
            flight.get("insurance", 0),
            profit,
            passenger.get("passport_number", "-"),
            "",  # cancelled
        ]
        ws.append_row(new_row, value_input_option="USER_ENTERED")
        print(f"[OK] Sheet updated for {company}: row {next_row}")
    except Exception as e:
        print(f"[ERROR] Sheet update failed: {e}")

def find_row_by_ref(company: str, ref: str):
    """پیدا کردن شماره ردیف در شیت بر اساس رفرنس — برمی‌گردونه (row_index, row_data) یا None"""
    try:
        ws = get_gsheet_ws(company)
        if not ws:
            return None
        all_rows = ws.get_all_values()
        for i, row in enumerate(all_rows):
            if len(row) >= 2 and row[COL_REF - 1].strip().upper() == ref.strip().upper():
                return (i + 1, row)  # 1-based
        return None
    except Exception as e:
        print(f"[ERROR] find_row_by_ref: {e}")
        return None

def sheet_update_cell(company: str, row_index: int, col: int, value):
    try:
        ws = get_gsheet_ws(company)
        if ws:
            ws.update_cell(row_index, col, value)
    except Exception as e:
        print(f"[ERROR] sheet_update_cell: {e}")

# ============================================================
# PDF GENERATION
# ============================================================
async def generate_pdf(passenger: dict, flight: dict, template_path: str):
    with open(template_path, 'r') as f:
        html = f.read()

    fname = passenger["first_name"]
    lname = passenger["last_name"]
    pp    = passenger["passport_number"]
    ref_out = gen_ref()
    ref_ret = gen_ref()

    is_oneway    = flight.get("type") == "oneway"
    flight_class = flight.get("class", "economy")
    outbound     = flight.get("outbound") or FLIGHTS["varesh_ika_tbs"]
    return_fl    = flight.get("return_flight") or FLIGHTS["varesh_tbs_ika"]

    baggage_key_out = f"{'qeshm' if 'Qeshm' in outbound['airline'] else 'varesh'}_{flight_class}"
    baggage_key_ret = f"{'qeshm' if 'Qeshm' in return_fl['airline'] else 'varesh'}_{flight_class}"
    baggage_out = BAGGAGE.get(baggage_key_out, "20 kg")
    baggage_ret = BAGGAGE.get(baggage_key_ret, "20 kg")

    sell_out = flight.get("sell_out", 0)
    sell_ret = flight.get("sell_ret", 0)
    date_out = flight.get("date_out", "")
    date_ret = flight.get("date_ret", "")

    out_src         = outbound.get("from", "IKA")
    original_out_src = out_src

    if is_oneway and out_src == "TBS":
        sell_ret = sell_out
        date_ret = date_out
        sell_out = 0
        date_out = ""
        return_fl = outbound

    if is_oneway:
        if original_out_src == "IKA":
            html = html.replace('.connector {', '.connector { display:none !important; } .flight-card:last-of-type { display:none !important; } .x_{')
        else:
            html = html.replace('.connector {', '.connector { display:none !important; } .flight-card:first-child { display:none !important; } .x_{')

    is_flyticket = 'FLYTICKET' in template_path.upper()

    if is_flyticket and is_oneway:
        if original_out_src == "IKA":
            html = re.sub(r'<hr class=\"leg-divider\">.*?</div><!-- card-body -->', '</div><!-- card-body -->', html, flags=re.DOTALL)
        else:
            html = re.sub(r'(<div class=\"card-body\">).*?<hr class=\"leg-divider\">', r'\1', html, flags=re.DOTALL)
        html = html.replace('<span>ROUND-TRIP</span>', '<span>ONE-WAY</span>')

    if is_flyticket:
        html = html.replace('FT_PASSENGER_NAME', f'{fname} {lname}')
        html = html.replace('FT_PASSPORT', pp)
    else:
        html = html.replace('MAHMOUD<br>ESLAMINOSRATABADI', f'{fname}<br>{lname}')
        html = html.replace('MAHMOUD ESLAMINOSRATABADI', f'{fname} {lname}')
        html = html.replace('B74341095', pp)

    if is_flyticket:
        html = html.replace('FT_REF_OUT', ref_out)
        html = html.replace('FT_REF_RET', ref_ret)
    else:
        html = html.replace('QA7X29', ref_out)
        html = html.replace('VR3M84', ref_ret)

    time_out = outbound.get('time', '---')
    time_ret = return_fl.get('time', '---')

    if is_flyticket:
        html = html.replace('FT_DATE_OUT', date_out if date_out else '---')
        html = html.replace('FT_DATE_RET', date_ret if date_ret else '---')
        html = html.replace('FT_TIME_OUT', time_out)
        html = html.replace('FT_TIME_RET', time_ret)
    else:
        html = html.replace('DATE_OUT', date_out if date_out else '---')
        html = html.replace('DATE_RET', date_ret if date_ret else '---')
        html = html.replace('FT_TIME_OUT', time_out)
        html = html.replace('FT_TIME_RET', time_ret)

    if is_flyticket:
        out_airline_short = 'QESHM' if 'Qeshm' in outbound['airline'] else 'VARESH'
        ret_airline_short = 'QESHM' if 'Qeshm' in return_fl['airline'] else 'VARESH'
        out_flight_num = outbound['flight'].split('.')[-1].strip()
        ret_flight_num = return_fl['flight'].split('.')[-1].strip()
        html = html.replace('>VARESH<', f'>{out_airline_short}<', 1)
        html = html.replace('>VARESH<', f'>{ret_airline_short}<', 1)
        html = html.replace('>6808<', f'>{out_flight_num}<', 1)
        html = html.replace('>6809<', f'>{ret_flight_num}<', 1)
    else:
        html = html.replace('Qeshm Airline', outbound['airline'])
        html = html.replace('QESHM AIRLINE', outbound['airline'].upper())
        html = html.replace('Q.2273', outbound['flight'])
        html = html.replace('Varesh Airline', return_fl['airline'])
        html = html.replace('VARESH AIRLINE', return_fl['airline'].upper())
        html = html.replace('VR.6809', return_fl['flight'])

    price_out_str = f"{sell_out:,}"
    price_ret_str = f"{sell_ret:,}"
    if is_flyticket:
        html = html.replace('174,300,000 IRR', f'{price_out_str} IRR', 1)
        html = html.replace('184,000,000 IRR', f'{price_ret_str} IRR', 1)
    else:
        html = html.replace('PRICE_OUT', f'{price_out_str}<br><span style="font-size:10px;font-weight:500;">تومان</span>')
        html = html.replace('PRICE_RET', f'{price_ret_str}<br><span style="font-size:10px;font-weight:500;">تومان</span>')

    if is_flyticket:
        html = html.replace('<strong>OK / 20 kg</strong>', f'<strong>OK / {baggage_out}</strong>', 1)
        html = html.replace('<strong>OK / 20 kg</strong>', f'<strong>OK / {baggage_ret}</strong>', 1)
    else:
        html = re.sub(r'<span class="ic-value">30 kg</span>', f'<span class="ic-value">{baggage_out}</span>', html, count=1)
        html = re.sub(r'<span class="ic-value">20 kg</span>', f'<span class="ic-value">{baggage_ret}</span>', html, count=1)

    if flight_class == "business":
        if is_flyticket:
            html = html.replace('<span>ECONOMY CLASS</span>', '<span>BUSINESS CLASS</span>')
        else:
            html = html.replace('>Economy<', '>Business<')

    if is_flyticket:
        compact_css = """
  <style>
    body { padding:0 !important; }
    .ticket { width:210mm !important; min-height:unset !important; height:297mm !important; padding:6mm 10mm !important; overflow:hidden !important; }
    .logo img { height:16mm !important; }
    .header { margin-bottom:2mm !important; }
    .divider, .divider-dashed { margin:1.5mm 0 !important; }
    .passenger-row { margin-bottom:0 !important; padding:1mm 0 !important; }
    .pax-block label { font-size:5.5pt !important; }
    .pax-block strong { font-size:9pt !important; }
    .card-header { padding:1.5mm 5mm !important; }
    .card-body { padding:0 7mm !important; }
    .route-row { margin-top:1.5mm !important; margin-bottom:0 !important; }
    .iata { font-size:22pt !important; }
    .apt-name { font-size:5.5pt !important; }
    .info-grid, .info-grid2 { padding:1.5mm 0 !important; }
    .info-cell label { font-size:5pt !important; margin-bottom:.5mm !important; }
    .info-cell strong { font-size:7.5pt !important; }
    .leg-divider { margin:0 5mm !important; }
    .sep1, .sep2 { margin:0 -7mm !important; }
    .notes-section { flex:1; display:flex; flex-direction:column; justify-content:space-between; min-height:0; }
    .notes-title { font-size:8pt !important; margin-bottom:1mm !important; }
    .note-item { font-size:5.5pt !important; line-height:1.4 !important; margin-bottom:0 !important; }
    .support-divider { margin:1mm 0 !important; }
    .support-title { font-size:6.5pt !important; margin-bottom:.5mm !important; }
    .support-item { font-size:6.5pt !important; line-height:1.5 !important; }
    .ad-box { margin-top:2mm !important; }
  </style>
"""
        html = html.replace('</head>', compact_css + '</head>')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page    = await browser.new_page()
        await page.set_content(html, wait_until='networkidle')
        await page.wait_for_timeout(1000)
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp_path = tmp.name
        await page.pdf(path=tmp_path, format='A4',
                       margin={'top':'0mm','bottom':'0mm','left':'0mm','right':'0mm'},
                       print_background=True)
        await browser.close()

    reader = PdfReader(tmp_path)
    writer = PdfWriter()
    for page_obj in reader.pages:
        writer.add_page(page_obj)
    writer.encrypt(user_password="", owner_password=PDF_PASSWORD, permissions_flag=4)

    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as locked:
        locked_path = locked.name
    with open(locked_path, 'wb') as f:
        writer.write(f)
    with open(locked_path, 'rb') as f:
        pdf_bytes = f.read()

    os.unlink(tmp_path)
    os.unlink(locked_path)
    return pdf_bytes, ref_out, ref_ret

# ============================================================
# EXCEL
# ============================================================
def update_excel(passenger: dict, flight: dict, ref_out: str, ref_ret: str):
    if not os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['بلیت‌ها']
    next_row = ws.max_row + 1
    light_blue = "EBF3FF"
    thin_gray  = Side(style="thin", color="CCCCCC")
    outbound  = flight.get("outbound") or {}
    return_fl = flight.get("return_flight") or {}
    is_round  = flight.get("type") == "roundtrip"
    total_sell = flight.get("sell_out", 0) + (flight.get("sell_ret", 0) if is_round else 0)
    total_buy  = flight.get("buy_out",  0) + (flight.get("buy_ret",  0) if is_round else 0)
    row_data = [
        next_row - 1, ref_out,
        f"{passenger['first_name']} {passenger['last_name']}",
        f"{outbound.get('from','-')} → {outbound.get('to','-')}",
        f"{return_fl.get('from','-')} → {return_fl.get('to','-')}" if is_round else "-",
        f"{outbound.get('airline','-')} {outbound.get('flight','')}",
        f"{return_fl.get('airline','-')} {return_fl.get('flight','')}" if is_round else "-",
        flight.get("date_out", "-"),
        flight.get("date_ret", "-") if is_round else "-",
        total_buy, total_sell,
        flight.get("voucher", 0), flight.get("insurance", 0),
        f"=(K{next_row}-J{next_row})+L{next_row}+M{next_row}",
        passenger.get("passport_number", "-"),
        "",
    ]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        cell.fill      = PatternFill("solid", start_color=light_blue)
    for col in [10, 11, 12, 13, 14]:
        ws.cell(row=next_row, column=col).number_format = '#,##0'
    wb.save(EXCEL_FILE)

# ============================================================
# ACCESS CONTROL
# ============================================================
ALLOWED_USERS = [178875046, 267020688]

def is_allowed(update) -> bool:
    return update.effective_user.id in ALLOWED_USERS

# ============================================================
# HANDLERS — صدور بلیت
# ============================================================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("❌ شما دسترسی به این ربات ندارید.")
        return ConversationHandler.END

    reply_keyboard = ReplyKeyboardMarkup([
        ["✈️ IRANFLY", "🎫 FLYTICKET"],
        ["🔍 Reference"],
        ["❌ لغو"]
    ], resize_keyboard=True)

    inline_keyboard = [
        [InlineKeyboardButton("✈️ IRANFLY",    callback_data="company_iranfly")],
        [InlineKeyboardButton("🎫 FLYTICKET",  callback_data="company_flyticket")],
        [InlineKeyboardButton("🔍 Reference",  callback_data="reference_menu")],
        [InlineKeyboardButton("❌ لغو",        callback_data="cancel_btn")],
    ]

    await update.message.reply_text(
        "✈️ *به سیستم صدور بلیت خوش آمدید!*\n\nلطفاً شرکت مورد نظر را انتخاب کنید:",
        parse_mode="Markdown",
        reply_markup=reply_keyboard
    )
    await update.message.reply_text("انتخاب کنید:", reply_markup=InlineKeyboardMarkup(inline_keyboard))
    return ConversationHandler.END

async def reply_keyboard_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return ConversationHandler.END
    text = update.message.text
    if text == "✈️ IRANFLY":
        context.user_data['company'] = 'iranfly'
        await update.message.reply_text("✅ IRANFLY انتخاب شد.\n\n📸 لطفاً عکس پاسپورت مسافر را ارسال کنید.")
        return WAITING_PASSPORT
    elif text == "🎫 FLYTICKET":
        context.user_data['company'] = 'flyticket'
        await update.message.reply_text("✅ FLYTICKET انتخاب شد.\n\n📸 لطفاً عکس پاسپورت مسافر را ارسال کنید.")
        return WAITING_PASSPORT
    elif text == "🔍 Reference":
        await update.message.reply_text("🔍 رفرنس بلیت را وارد کنید:")
        return WAITING_REF_INPUT
    elif text == "❌ لغو":
        context.user_data.clear()
        await update.message.reply_text("❌ لغو شد.")
        return ConversationHandler.END

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "company_iranfly":
        context.user_data['company'] = 'iranfly'
        await query.message.reply_text("✅ IRANFLY انتخاب شد.\n\n📸 لطفاً عکس پاسپورت مسافر را ارسال کنید.")
        return WAITING_PASSPORT
    elif query.data == "company_flyticket":
        context.user_data['company'] = 'flyticket'
        await query.message.reply_text("✅ FLYTICKET انتخاب شد.\n\n📸 لطفاً عکس پاسپورت مسافر را ارسال کنید.")
        return WAITING_PASSPORT
    elif query.data == "reference_menu":
        await query.message.reply_text("🔍 رفرنس بلیت را وارد کنید:")
        return WAITING_REF_INPUT
    elif query.data == "cancel_btn":
        context.user_data.clear()
        await query.message.reply_text("❌ لغو شد.")
        return ConversationHandler.END

    # ویرایش فیلد
    elif query.data.startswith("edit_"):
        field = query.data[5:]
        context.user_data['edit_field'] = field
        prompts = {
            "buy":       "💰 قیمت خرید جدید را وارد کنید (عدد):",
            "sell":      "💵 قیمت فروش جدید را وارد کنید (عدد):",
            "passenger": "👤 نام و فامیل جدید مسافر را وارد کنید:",
            "passport":  "🔢 شماره پاسپورت جدید را وارد کنید:",
            "cancel":    "❌ آیا مطمئنید که می‌خواهید این بلیت را باطل کنید؟",
        }
        if field == "cancel":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ بله، باطل شود", callback_data="confirm_cancel")],
                [InlineKeyboardButton("🔙 خیر", callback_data="back_to_edit")],
            ])
            await query.message.reply_text(prompts[field], reply_markup=kb)
            return WAITING_CANCEL_CONFIRM
        else:
            await query.message.reply_text(prompts.get(field, "مقدار جدید را وارد کنید:"))
            return WAITING_EDIT_VALUE

    elif query.data == "confirm_cancel":
        await query.message.reply_text("💵 مبلغ سود باطلی را وارد کنید (عدد):")
        context.user_data['edit_field'] = 'do_cancel'
        return WAITING_EDIT_VALUE

    elif query.data == "back_to_edit":
        await _show_edit_menu(query.message, context)
        return WAITING_EDIT_VALUE

async def _show_edit_menu(message, context: ContextTypes.DEFAULT_TYPE):
    ref  = context.user_data.get('edit_ref', '')
    row  = context.user_data.get('edit_row_data', [])
    name = row[COL_PASSENGER - 1] if len(row) >= COL_PASSENGER else '-'
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("💰 قیمت خرید",     callback_data="edit_buy")],
        [InlineKeyboardButton("💵 قیمت فروش",     callback_data="edit_sell")],
        [InlineKeyboardButton("👤 مسافر",          callback_data="edit_passenger")],
        [InlineKeyboardButton("🔢 شماره پاسپورت", callback_data="edit_passport")],
        [InlineKeyboardButton("❌ باطل کردن",      callback_data="edit_cancel")],
    ])
    await message.reply_text(
        f"🔍 رفرنس: `{ref}`\n👤 مسافر: {name}\n\nکدام فیلد را ویرایش می‌کنید؟",
        parse_mode="Markdown",
        reply_markup=kb
    )

# ============================================================
# HANDLERS — Reference جستجو و ویرایش
# ============================================================
async def receive_ref_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return ConversationHandler.END
    ref     = update.message.text.strip().upper()
    company = context.user_data.get('company', 'iranfly')

    result = find_row_by_ref(company, ref)
    if not result:
        # شرکت دیگه رو هم چک کن
        other = 'flyticket' if company == 'iranfly' else 'iranfly'
        result = find_row_by_ref(other, ref)
        if result:
            company = other
            context.user_data['company'] = company

    if not result:
        await update.message.reply_text(f"❌ رفرنس `{ref}` پیدا نشد.", parse_mode="Markdown")
        return WAITING_REF_INPUT

    row_index, row_data = result
    context.user_data['edit_ref']      = ref
    context.user_data['edit_row_index'] = row_index
    context.user_data['edit_row_data']  = row_data

    await _show_edit_menu(update.message, context)
    return WAITING_EDIT_VALUE

async def receive_edit_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        return ConversationHandler.END

    value     = update.message.text.strip()
    field     = context.user_data.get('edit_field')
    company   = context.user_data.get('company', 'iranfly')
    row_index = context.user_data.get('edit_row_index')
    row_data  = context.user_data.get('edit_row_data', [])
    ref       = context.user_data.get('edit_ref', '')

    if not row_index:
        await update.message.reply_text("❌ خطا: ردیف پیدا نشد.")
        return ConversationHandler.END

    try:
        if field == "buy":
            num = to_num(value)
            sheet_update_cell(company, row_index, COL_BUY, num)
            await update.message.reply_text(f"✅ قیمت خرید به `{num:,}` تومان آپدیت شد.", parse_mode="Markdown")

        elif field == "sell":
            num = to_num(value)
            sheet_update_cell(company, row_index, COL_SELL, num)
            await update.message.reply_text(f"✅ قیمت فروش به `{num:,}` تومان آپدیت شد.", parse_mode="Markdown")

        elif field == "passenger":
            parts = value.split()
            if len(parts) < 2:
                await update.message.reply_text("❌ لطفاً نام و فامیل را با فاصله وارد کنید.")
                return WAITING_EDIT_VALUE
            sheet_update_cell(company, row_index, COL_PASSENGER, value.upper())

            # صدور بلیت جدید
            passport_num = row_data[COL_PASSPORT - 1] if len(row_data) >= COL_PASSPORT else "UNKNOWN"
            passenger = {
                "first_name":      parts[0].upper(),
                "last_name":       " ".join(parts[1:]).upper(),
                "passport_number": passport_num,
            }
            flight = _rebuild_flight_from_row(row_data)
            template = FLYTICKET_TEMPLATE_FILE if company == 'flyticket' else TEMPLATE_FILE
            await update.message.reply_text("⏳ در حال صدور بلیت جدید...")
            pdf_bytes, new_ref, _ = await generate_pdf(passenger, flight, template)
            filename = f"{passenger['first_name']}_{passenger['last_name']}.pdf"
            await update.message.reply_document(
                document=pdf_bytes, filename=filename,
                caption=f"✅ مسافر آپدیت شد.\n🔖 رفرنس جدید: `{new_ref}`",
                parse_mode="Markdown"
            )

        elif field == "passport":
            sheet_update_cell(company, row_index, COL_PASSPORT, value.upper())

            # صدور بلیت جدید
            full_name = row_data[COL_PASSENGER - 1] if len(row_data) >= COL_PASSENGER else "UNKNOWN PASSENGER"
            name_parts = full_name.split()
            passenger = {
                "first_name":      name_parts[0] if name_parts else "UNKNOWN",
                "last_name":       " ".join(name_parts[1:]) if len(name_parts) > 1 else "PASSENGER",
                "passport_number": value.upper(),
            }
            flight = _rebuild_flight_from_row(row_data)
            template = FLYTICKET_TEMPLATE_FILE if company == 'flyticket' else TEMPLATE_FILE
            await update.message.reply_text("⏳ در حال صدور بلیت جدید...")
            pdf_bytes, new_ref, _ = await generate_pdf(passenger, flight, template)
            filename = f"{passenger['first_name']}_{passenger['last_name']}.pdf"
            await update.message.reply_document(
                document=pdf_bytes, filename=filename,
                caption=f"✅ پاسپورت آپدیت شد.\n🔖 رفرنس جدید: `{new_ref}`",
                parse_mode="Markdown"
            )

        elif field == "do_cancel":
            profit = to_num(value)
            sheet_update_cell(company, row_index, COL_BUY,       0)
            sheet_update_cell(company, row_index, COL_SELL,      0)
            sheet_update_cell(company, row_index, COL_PROFIT,    profit)
            sheet_update_cell(company, row_index, COL_CANCELLED, "✅ باطل")
            await update.message.reply_text(
                f"✅ بلیت `{ref}` باطل شد.\n💵 سود ثبت‌شده: `{profit:,}` تومان",
                parse_mode="Markdown"
            )

    except Exception as e:
        await update.message.reply_text(f"❌ خطا: {str(e)}")

    context.user_data.clear()
    return ConversationHandler.END

def _rebuild_flight_from_row(row: list) -> dict:
    """بازسازی دیکشنری flight از ردیف گوگل شیت"""
    def safe(i):
        return row[i] if len(row) > i else "-"

    route_out = safe(COL_ROUTE_OUT - 1)   # e.g. "IKA → TBS"
    route_ret = safe(COL_ROUTE_RET - 1)
    air_out   = safe(COL_AIR_OUT - 1)     # e.g. "Varesh Airline VR.6808"
    air_ret   = safe(COL_AIR_RET - 1)
    date_out  = safe(COL_DATE_OUT - 1)
    date_ret  = safe(COL_DATE_RET - 1)
    sell_out  = int(str(safe(COL_SELL - 1)).replace(',', '') or 0)
    buy_out   = int(str(safe(COL_BUY  - 1)).replace(',', '') or 0)

    def parse_route(r):
        parts = [p.strip() for p in r.split('→')]
        return parts[0] if len(parts) > 0 else "IKA", parts[1] if len(parts) > 1 else "TBS"

    src_out, dst_out = parse_route(route_out)
    src_ret, dst_ret = parse_route(route_ret)

    def find_flight(airline_str, src, dst):
        a = "qeshm" if "qeshm" in airline_str.lower() else "varesh"
        key = f"{a}_{src.lower()}_{dst.lower()}"
        return FLIGHTS.get(key, FLIGHTS["varesh_ika_tbs"])

    outbound  = find_flight(air_out, src_out, dst_out)
    return_fl = find_flight(air_ret, src_ret, dst_ret)

    is_round = route_ret and route_ret != "-"
    flight_class = "economy"
    for f in FLIGHTS.values():
        if f.get("flight", "") in air_out:
            break

    return {
        "type":          "roundtrip" if is_round else "oneway",
        "outbound":      outbound,
        "return_flight": return_fl if is_round else None,
        "sell_out":      sell_out,
        "buy_out":       buy_out,
        "sell_ret":      0,
        "buy_ret":       0,
        "voucher":       0,
        "insurance":     0,
        "class":         flight_class,
        "date_out":      date_out,
        "date_ret":      date_ret if is_round else "",
    }

# ============================================================
# HANDLERS — پاسپورت و پرواز
# ============================================================
async def receive_passport(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("❌ شما دسترسی به این ربات ندارید.")
        return ConversationHandler.END
    await update.message.reply_text("📷 پاسپورت دریافت شد. در حال استخراج اطلاعات...")
    photo      = update.message.photo[-1]
    file       = await context.bot.get_file(photo.file_id)
    image_bytes = await file.download_as_bytearray()
    try:
        passport_info = extract_passport_info(bytes(image_bytes))
        context.user_data['passport'] = passport_info
        await update.message.reply_text(
            f"✅ اطلاعات استخراج شد:\n"
            f"👤 نام: {passport_info['first_name']} {passport_info['last_name']}\n"
            f"🔢 پاسپورت: {passport_info['passport_number']}\n\n"
            "حالا اطلاعات پرواز را وارد کنید:\n\n"
            "فرمت:\n`ika/tbs/qeshm/6jun/sell14.000/buy12.000`\n"
            "دو طرفه: خط اول رفت، خط دوم برگشت",
            parse_mode="Markdown"
        )
        return WAITING_FLIGHT_INFO
    except Exception as e:
        await update.message.reply_text(f"❌ خطا در استخراج اطلاعات: {str(e)}")
        return WAITING_PASSPORT

async def receive_flight_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text     = update.message.text
    passport = context.user_data.get('passport')
    if not passport:
        await update.message.reply_text("❌ ابتدا عکس پاسپورت ارسال کنید.")
        return WAITING_PASSPORT
    await update.message.reply_text("⏳ در حال صدور بلیت...")
    try:
        company  = context.user_data.get('company', 'iranfly')
        template = FLYTICKET_TEMPLATE_FILE if company == 'flyticket' else TEMPLATE_FILE
        flight   = parse_flight_input(text)
        pdf_bytes, ref_out, ref_ret = await generate_pdf(passport, flight, template)
        filename = f"{passport['first_name']}_{passport['last_name']}.pdf"
        await update.message.reply_document(
            document=pdf_bytes, filename=filename,
            caption=f"🔖 رفرنس رفت: `{ref_out}`" + (f"\n🔖 رفرنس برگشت: `{ref_ret}`" if flight['type'] == 'roundtrip' else ""),
            parse_mode="Markdown"
        )
        update_excel(passport, flight, ref_out, ref_ret)
        update_google_sheet(company, passport, flight, ref_out, ref_ret)
        await update.message.reply_text(
            "مسافر محترم، مدارک شما در حال صدور بوده و در اسرع وقت ارسال خواهد شد. "
            "ممنون از صبر و شکیبایی شما. 🙏"
        )
        context.user_data.clear()
        return ConversationHandler.END
    except Exception as e:
        await update.message.reply_text(f"❌ خطا: {str(e)}\nلطفاً دوباره تلاش کنید.")
        return WAITING_FLIGHT_INFO

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("❌ لغو شد.")
    return ConversationHandler.END

# ============================================================
# MAIN
# ============================================================
def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CallbackQueryHandler(button_handler))

    conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(filters.PHOTO, receive_passport),
            MessageHandler(filters.Text(["✈️ IRANFLY", "🎫 FLYTICKET", "🔍 Reference", "❌ لغو"]), reply_keyboard_handler),
        ],
        states={
            WAITING_PASSPORT:       [MessageHandler(filters.PHOTO, receive_passport)],
            WAITING_FLIGHT_INFO:    [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_flight_info)],
            WAITING_REF_INPUT:      [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_ref_input)],
            WAITING_EDIT_VALUE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_edit_value)],
            WAITING_CANCEL_CONFIRM: [],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            MessageHandler(filters.Text(["❌ لغو"]), reply_keyboard_handler),
        ],
    )
    app.add_handler(conv)
    print("🚀 IRANFLY Bot started...")
    app.run_polling()

if __name__ == "__main__":
    main()
